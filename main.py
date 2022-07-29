import datetime
import itertools

from openpyxl import Workbook, load_workbook
from sqlalchemy.sql import extract
from telegram import ReplyKeyboardMarkup
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler

from data import db_session
from data.brends import Brends
from data.delivery_goods import Delivery_goods
from data.deliverymen import Deliverymen
from data.goods import Goods
from data.purchase import Purchase
from data.sales import Sales
from settings import add_deliverymen
from settings import admin, admin_2
from settings import deliverymen
from settings import delyverymen_id
from settings import sberbank, alfabank
from settings import text_chat
from settings import text_start

TOKEN = '5478798012:AAHOWISNLFiv4P4qusEsnoJgJySTADiSnak'


def start(update, context):
    try:
        context.user_data['locality'] = {}
        context.user_data['locality'][1] = 'Старт'
        if update.message.chat.id == admin or update.message.chat.id == admin_2:
            update.message.reply_text('Нажмите кнопки снизу',
                                      reply_markup=ReplyKeyboardMarkup([['Наличие', 'Изменить количество'],
                                                                        ['Добавить линейку', 'Изменить линейку'],
                                                                        ['Проверка', 'Статистика'],
                                                                        ['Получить форму', 'Выслать форму'],
                                                                        ['Новый закуп', 'Добавить описание'],
                                                                        ['Описание вкусов']],
                                                                       resize_keyboard=True,
                                                                       one_time_keyboard=True))
        elif update.message.chat.id in deliverymen.values():
            context.user_data['user_id_db'] = db_session.create_session().query(
                Deliverymen).filter(Deliverymen.user_id == update.message.chat.id).first().id
            update.message.reply_text('Нажмите кнопки снизу', reply_markup=ReplyKeyboardMarkup([['Продать', 'Наличие'],
                                                                                                ['Статистика за день',
                                                                                                 'Статистика за месяц'],
                                                                                                [
                                                                                                    'Отправить на проверку',
                                                                                                    'Нужно перевести'],
                                                                                                ['Описание вкусов']],
                                                                                               resize_keyboard=True,
                                                                                               one_time_keyboard=True))
        else:
            update.message.reply_text(text_start)
            update.message.reply_text('Нажмите кнопки на клавиатуре',
                                      reply_markup=ReplyKeyboardMarkup([['Наличие'],
                                                                        ['Описание вкусов'],
                                                                        ['Доставка']],
                                                                       resize_keyboard=True,
                                                                       one_time_keyboard=True))
    except:
        return error_handler(update, context)


def error_handler(update, context):
    update.message.reply_text('Упс... Что-то пошло не так')
    context.user_data['locality'] = {}
    context.user_data['locality'][1] = 'Старт'
    if update.message.chat.id == admin or update.message.chat.id == admin_2:
        update.message.reply_text('Нажмите кнопки снизу',
                                  reply_markup=ReplyKeyboardMarkup([['Наличие', 'Изменить количество'],
                                                                    ['Добавить линейку', 'Изменить линейку'],
                                                                    ['Проверка', 'Статистика'],
                                                                    ['Получить форму', 'Выслать форму'],
                                                                    ['Новый закуп', 'Добавить описание'],
                                                                    ['Описание вкусов']],
                                                                   resize_keyboard=True,
                                                                   one_time_keyboard=True))
    elif update.message.chat.id in deliverymen.values():
        context.user_data['user_id_db'] = db_session.create_session().query(
            Deliverymen).filter(Deliverymen.user_id == update.message.chat.id).first().id
        update.message.reply_text('Нажмите кнопки снизу', reply_markup=ReplyKeyboardMarkup([['Продать', 'Наличие'],
                                                                                            ['Статистика за день',
                                                                                             'Статистика за месяц'],
                                                                                            [
                                                                                                'Отправить на проверку',
                                                                                                'Нужно перевести'],
                                                                                            ['Описание вкусов']],
                                                                                           resize_keyboard=True,
                                                                                           one_time_keyboard=True))
    else:
        update.message.reply_text('Нажмите кнопки на клавиатуре',
                                  reply_markup=ReplyKeyboardMarkup([['Наличие'],
                                                                    ['Описание вкусов'],
                                                                    ['Доставка']],
                                                                   resize_keyboard=True,
                                                                   one_time_keyboard=True))


def start_menu_handler(update, context):
    try:
        context.user_data['locality'] = {}
        context.user_data['locality'][1] = 'Старт'
        if update.message.chat.id == admin or update.message.chat.id == admin_2:
            update.message.reply_text('Возврат в меню',
                                      reply_markup=ReplyKeyboardMarkup([['Наличие', 'Изменить количество'],
                                                                        ['Добавить линейку', 'Изменить линейку'],
                                                                        ['Проверка', 'Статистика'],
                                                                        ['Получить форму', 'Выслать форму'],
                                                                        ['Новый закуп', 'Добавить описание'],
                                                                        ['Описание вкусов']],
                                                                       resize_keyboard=True,
                                                                       one_time_keyboard=True))
        elif update.message.chat.id in deliverymen.values():
            context.user_data['user_id_db'] = db_session.create_session().query(
                Deliverymen).filter(Deliverymen.user_id == update.message.chat.id).first().id
            update.message.reply_text('Возврат в меню', reply_markup=ReplyKeyboardMarkup([['Продать', 'Наличие'],
                                                                                          ['Статистика за день',
                                                                                           'Статистика за месяц'],
                                                                                          [
                                                                                              'Отправить на проверку',
                                                                                              'Нужно перевести'],
                                                                                          ['Описание вкусов']],
                                                                                         resize_keyboard=True,
                                                                                         one_time_keyboard=True))
        else:
            update.message.reply_text('Возврат в меню',
                                      reply_markup=ReplyKeyboardMarkup([['Наличие'],
                                                                        ['Описание вкусов'],
                                                                        ['Доставка']],
                                                                       resize_keyboard=True,
                                                                       one_time_keyboard=True))
    except:
        return error_handler(update, context)


def handler(update, context):
    if update.message.chat.id == admin or update.message.chat.id == admin_2:
        # Главное меню__
        if context.user_data['locality'][len(context.user_data['locality'])] == 'Старт':
            if update.message.text == 'Добавить линейку':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Добавить линейку'
                update.message.reply_text(f'Введите название, цену, зарплату скидки '
                                          f'через слэш каждый с новой строки',
                                          reply_markup=ReplyKeyboardMarkup([['Назад']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Изменить линейку':
                reply_keyboard = [[elem.brend] for elem in sorted(db_session.create_session().query(Brends).all(),
                                                                  key=lambda x: -(x.price))]
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = \
                    'Изменить линейку'
                update.message.reply_text('Выберите линейку для изменения',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Изменить количество':
                reply_keyboard = [[elem] for elem in deliverymen]
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][
                    len(context.user_data['locality']) + 1] = 'Изменить количество доставщика'
                update.message.reply_text('Выберите доставщика для изменения',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Наличие':
                reply_keyboard = [[elem] for elem in deliverymen]
                reply_keyboard.append(['Общее'])
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Наличие'
                update.message.reply_text('Выберите доставщика',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Проверка':
                db_sess = db_session.create_session()
                send_on_check = db_sess.query(Sales).filter(Sales.on_check == True, Sales.is_send == False).all()
                if send_on_check:
                    context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Проверка 1'
                    reply_keyboard = []
                    for elem in send_on_check:
                        reply_keyboard.append([str(elem.date) + ' ' +
                                               str(db_sess.query(
                                                   Deliverymen).get(elem.deliveryman_id).name) + ' ' +
                                               str(elem.total)])
                    reply_keyboard.append(['Отмена'])
                    update.message.reply_text('Выберите дату',
                                              reply_markup=ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True,
                                                                               one_time_keyboard=True))
                else:
                    update.message.reply_text('Ничего не найдено')
                    return start_menu_handler(update, context)
            elif update.message.text == 'Статистика':
                reply_keyboard = [[elem] for elem in deliverymen]
                reply_keyboard.append(['Общее'])
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Статистика'
                update.message.reply_text('Выберите доставщика',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Получить форму':
                db_sess = db_session.create_session()
                wb = Workbook()
                ws = wb.active
                ws.append(['', ''] + list(deliverymen.keys()))
                for elem in db_sess.query(Brends).all():
                    ws.append([elem.brend])
                    for ele in db_sess.query(Goods).filter(Goods.brend == elem).all():
                        ws.append(['', ele.title])
                wb.save('form.xlsx')
                update.message.reply_document(document=open('form.xlsx', 'rb'), filename='form.xlsx')
                return start_menu_handler(update, context)
            elif update.message.text == 'Выслать форму':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Выслать форму 2'
                update.message.reply_text('Пришлите файл',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Новый закуп':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Новый закуп 2'
                update.message.reply_text('Пришлите файл',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Добавить описание':
                reply_keyboard = [[elem.brend] for elem in sorted(db_session.create_session().query(Brends).all(),
                                                                  key=lambda x: -(x.price))]
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Добавить описание 2'
                update.message.reply_text('Выберите линейку для добавления описания',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Описание вкусов':
                reply_keyboard = [[elem.brend] for elem in sorted(db_session.create_session(
                ).query(Brends).filter(Brends.photo_link != '').all(), key=lambda x: -(x.price))]
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Описание вкусов 2'
                update.message.reply_text('Выберите линейку',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                error_handler(update, context)
        #
        # Добавление новой линейки
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Добавить линейку':
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            else:
                context.user_data['new_good'] = {0: update.message.text.split('\n')[0]}
                db_sess = db_session.create_session()
                db_sess.add(Brends(brend=context.user_data['new_good'][0].rstrip(),
                                   price=int(update.message.text.split('\n')[1]),
                                   discount_1=int(update.message.text.split('\n')[3].split('/')[0]),
                                   discount_2=int(update.message.text.split('\n')[3].split('/')[1]),
                                   salary=int(update.message.text.split('\n')[2])))
                db_sess.commit()
                context.user_data['locality'][
                    len(context.user_data['locality']) + 1] = 'Добавить вкус новой линейки'
                update.message.reply_text(
                    f'Введите вкусы {context.user_data["new_good"][0]}, каждый с новой строки',
                    reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                     resize_keyboard=True,
                                                     one_time_keyboard=True))
        elif context.user_data['locality'][len(context.user_data['locality'])] == \
                'Добавить вкус новой линейки':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            else:
                if add_goods(context.user_data['new_good'][0], update.message.text.split('\n')):
                    update.message.reply_text('Бренд и вкусы успешно добавлены')
                    return start_menu_handler(update, context)
                else:
                    return error_handler(update, context)
        #
        # Изменение линейки
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Изменить линейку':
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            else:
                brend_id = check_brend(update.message.text)
                if brend_id:
                    brend = db_session.create_session().query(Brends).get(brend_id)
                    context.user_data['redactor_brend'] = {0: brend_id}
                    context.user_data['locality'][
                        len(context.user_data['locality']) + 1] = 'Выбор изменения в линейке'
                    update.message.reply_text('Что вы хотите изменить?',
                                              reply_markup=ReplyKeyboardMarkup([[f'Цену {brend.price}'],
                                                                                [f'Название {brend.brend}'],
                                                                                [f'Зарплату {brend.salary}'],
                                                                                [f'Скидку {brend.discount_1}/'
                                                                                 f'{brend.discount_2}'],
                                                                                ['Изменить вкус'],
                                                                                ['Добавить вкус'],
                                                                                ['Удалить линейку'],
                                                                                ['Отмена']],
                                                                               resize_keyboard=True,
                                                                               one_time_keyboard=True))
                else:
                    update.message.reply_text('Линейка не найдена')
                    return start_menu_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Выбор изменения в линейке':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif update.message.text.split(' ')[0] == 'Цену':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Изменение цены в линейке'
                update.message.reply_text('Введите новую цену',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text.split(' ')[0] == 'Название':
                context.user_data['locality'][
                    len(context.user_data['locality']) + 1] = 'Изменение названия в линейке'
                update.message.reply_text('Введите новое название',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text.split(' ')[0] == 'Зарплату':
                context.user_data['locality'][
                    len(context.user_data['locality']) + 1] = 'Изменение зарплаты в линейке'
                update.message.reply_text('Введите новую зарплату',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Изменить вкус':
                reply_keyboard = [[elem.title] for elem in db_session.create_session().query(
                    Goods.title).filter(
                    Goods.brend_id == context.user_data['redactor_brend'][0]).all()]
                reply_keyboard.append(['Отмена'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Изменить вкус'
                update.message.reply_text('Выберите вкус для изменения',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Добавить вкус':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Добавить вкус'
                update.message.reply_text('Введите название вкуса',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text.split(' ')[0] == 'Скидку':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Изменить скидку'
                update.message.reply_text('Введите новую скидку через /',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Удалить линейку':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Удалить линейку'
                update.message.reply_text('Вы уверены?',
                                          reply_markup=ReplyKeyboardMarkup([['Да'],
                                                                            ['Нет'],
                                                                            ['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                update.message.reply_text('Нажмите кнопку',
                                          reply_markup=ReplyKeyboardMarkup([['Цену'],
                                                                            ['Название'],
                                                                            ['Изменить вкус'],
                                                                            ['Добавить вкус'],
                                                                            ['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Изменение цены в линейке':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif redact_brend_price(context.user_data['redactor_brend'][0], int(update.message.text)):
                update.message.reply_text('Цена успешно изменена')
                return start_menu_handler(update, context)
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Изменение названия в линейке':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif redact_brend_title(context.user_data['redactor_brend'][0], update.message.text):
                update.message.reply_text('Название успешно изменено')
                return start_menu_handler(update, context)
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Изменение зарплаты в линейке':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif redact_brend_salary(context.user_data['redactor_brend'][0], update.message.text):
                update.message.reply_text('Зарплата успешно изменена')
                return start_menu_handler(update, context)
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Изменить вкус':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            else:
                context.user_data['redactor_brend'][1] = update.message.text
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Новое название вкуса'
                update.message.reply_text('Введите новое название вкуса',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Добавить вкус':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            else:
                if add_good(context.user_data['redactor_brend'][0], update.message.text):
                    update.message.reply_text('Вкус успешно добавлен')
                    return start_menu_handler(update, context)
                else:
                    return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Новое название вкуса':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif redact_good_title(context.user_data['redactor_brend'][0], context.user_data['redactor_brend'][1],
                                   update.message.text):
                update.message.reply_text('Название успешно изменено')
                return start_menu_handler(update, context)
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Изменить скидку':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif '/' in update.message.text:
                db_sess = db_session.create_session()
                brend = db_sess.query(Brends).get(context.user_data['redactor_brend'][0])
                brend.discount_1 = int(update.message.text.split('/')[0])
                brend.discount_2 = int(update.message.text.split('/')[1])
                db_sess.add(brend)
                db_sess.commit()
                update.message.reply_text('Успешно')
                return start_menu_handler(update, context)
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Удалить линейку':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif update.message.text == 'Да':
                if delete_brand(context.user_data['redactor_brend'][0]):
                    update.message.reply_text('Успешно')
                    return start_menu_handler(update, context)
                else:
                    return error_handler(update, context)
            elif update.message.text == 'Нет':
                return start_menu_handler(update, context)
            else:
                return error_handler(update, context)
        #
        # Изменение количества у доставщика
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Изменить количество доставщика':
            context.user_data['add_amount'] = {'delivery_good_id': None, 'brend_id': None,
                                               'good_id': None, 'deliveryman_id': None, 'amount': None}
            db_sess = db_session.create_session()
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            elif (update.message.text,) in db_sess.query(Deliverymen.name).all():
                context.user_data['add_amount']['deliveryman_id'] = db_sess.query(
                    Deliverymen).filter(Deliverymen.name == update.message.text).first().id
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Выбор изменения доставщика'
                update.message.reply_text('Выберите, что хотите сделать',
                                          reply_markup=ReplyKeyboardMarkup([['Убавить товар'],
                                                                            ['Добавить товар'],
                                                                            ['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Выбор изменения доставщика':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif update.message.text == 'Убавить товар':
                reply_keyboard = [[elem.brend] for elem in db_session.create_session().query(
                    Brends).all()]
                reply_keyboard.append(['Отмена'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Убавить товар'
                update.message.reply_text('Выберите линейку для изменения',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Добавить товар':
                reply_keyboard = [[elem.brend] for elem in db_session.create_session().query(Brends).all()]
                reply_keyboard.append(['Отмена'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Добавить товар'
                update.message.reply_text('Выберите линейку для изменения',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Убавить товар':
            db_sess = db_session.create_session()
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif (update.message.text,) in db_sess.query(Brends.brend).all():
                brend = db_sess.query(Brends).filter(Brends.brend == update.message.text).first()
                context.user_data['add_amount']['brend_id'] = brend.id
                reply_keyboard = [[f"{elem.title} - "
                                   f"{get_amount(db_sess, elem, context)}"]
                                  for elem in db_sess.query(Goods).filter(Goods.brend == brend).all()]
                reply_keyboard.append(['Отмена'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = \
                    'Выбрать товар доставщика убавления'
                update.message.reply_text('Выберите товар линейки',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Добавить товар':
            db_sess = db_session.create_session()
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif (update.message.text,) in db_sess.query(Brends.brend).all():
                brend = db_sess.query(Brends).filter(Brends.brend == update.message.text).first()
                context.user_data['add_amount']['brend_id'] = brend.id
                reply_keyboard = [[f"{elem.title} - "
                                   f"{get_amount(db_sess, elem, context)}"]
                                  for elem in db_sess.query(Goods).filter(Goods.brend == brend).all()]
                reply_keyboard.append(['Отмена'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = \
                    'Выбрать товар доставщика добавления'
                update.message.reply_text('Выберите товар линейки',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == \
                'Выбрать товар доставщика убавления':
            db_sess = db_session.create_session()
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif (update.message.text.split(' -')[0],) in db_sess.query(Goods.title).all():
                good_deliver = db_sess.query(Delivery_goods).filter(
                    Delivery_goods.good_id == db_sess.query(
                        Goods
                    ).filter(
                        Goods.title == update.message.text.split(' -')[0],
                        Goods.brend_id == context.user_data['add_amount']['brend_id']
                    ).first().id,
                    Delivery_goods.deliveryman_id == context.user_data['add_amount']['deliveryman_id']).first()
                context.user_data['add_amount']['delivery_good_id'] = good_deliver.id
                context.user_data['add_amount']['amount'] = good_deliver.amount
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Ввод количества убавления'
                update.message.reply_text(f"Введите количество товара"
                                          f" меньше {context.user_data['add_amount']['amount']}",
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == \
                'Выбрать товар доставщика добавления':
            db_sess = db_session.create_session()
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif (update.message.text.split(' -')[0],) in db_sess.query(Goods.title).all():
                good_deliver = db_sess.query(Delivery_goods).filter(
                    Delivery_goods.good_id == db_sess.query(
                        Goods
                    ).filter(
                        Goods.title == update.message.text.split(' -')[0],
                        Goods.brend_id == context.user_data['add_amount']['brend_id']
                    ).first().id,
                    Delivery_goods.deliveryman_id ==
                    context.user_data['add_amount']['deliveryman_id']).first()
                context.user_data['add_amount']['delivery_good_id'] = good_deliver.id
                context.user_data['add_amount']['amount'] = good_deliver.amount
                context.user_data['locality'][len(context.user_data['locality']) + 1] = \
                    'Ввод количества добавления'
                update.message.reply_text('Введите количество товара',
                                          reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == \
                'Ввод количества убавления':
            try:
                if update.message.text == 'Отмена':
                    return start_menu_handler(update, context)
                elif 0 < int(update.message.text) <= context.user_data['add_amount']['amount']:
                    db_sess = db_session.create_session()
                    good_deliver = db_sess.query(Delivery_goods).filter(Delivery_goods.id ==
                                                                        context.user_data['add_amount']
                                                                        ['delivery_good_id']).first()
                    good_deliver.amount -= int(update.message.text)
                    db_sess.add(good_deliver)
                    brend_purch = db_sess.query(Purchase
                                                ).filter(Purchase.amount > 0,
                                                         Purchase.brend_id == context.user_data['add_amount'
                                                         ]['brend_id'
                                                         ]).first()
                    brend_purch.amount -= int(update.message.text)
                    db_sess.add(brend_purch)
                    db_sess.commit()
                    context.user_data['locality'][len(context.user_data['locality']) + 1] = \
                        'Согласие на возврат к вкусам'
                    update.message.reply_text('Хотите ещё изменить количество у других вкусов?',
                                              reply_markup=ReplyKeyboardMarkup([['Да'], ['Нет']],
                                                                               resize_keyboard=True,
                                                                               one_time_keyboard=True))
                else:
                    update.message.reply_text(f'Введённое количество превышает количество товара или'
                                              f' меньше 0. Введите его снова')
            except:
                update.message.reply_text(f'Ошибка. Введите еще раз')
        elif context.user_data['locality'][len(context.user_data['locality'])] == \
                'Ввод количества добавления':
            try:
                if update.message.text == 'Отмена':
                    return start_menu_handler(update, context)
                else:
                    db_sess = db_session.create_session()
                    good_deliver = db_sess.query(Delivery_goods).filter(Delivery_goods.id ==
                                                                        context.user_data['add_amount']
                                                                        ['delivery_good_id']).first()
                    good_deliver.amount += int(update.message.text)
                    db_sess.add(good_deliver)
                    brend_purch = db_sess.query(Purchase
                                                ).filter(Purchase.amount > 0,
                                                         Purchase.brend_id == context.user_data['add_amount'
                                                         ]['brend_id'
                                                         ]).all()[-1]
                    brend_purch.amount += int(update.message.text)
                    db_sess.add(brend_purch)
                    db_sess.commit()
                    context.user_data['locality'][len(context.user_data['locality']) + 1] = \
                        'Согласие на возврат к вкусам'
                    update.message.reply_text(f'Успешно. Хотите проделать то'
                                              f' же действие с другими вкусами?',
                                              reply_markup=ReplyKeyboardMarkup([['Да'], ['Нет']],
                                                                               resize_keyboard=True,
                                                                               one_time_keyboard=True))
            except:
                update.message.reply_text(f"Ошибка. Введите еще раз")
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Согласие на возврат к вкусам':
            if update.message.text == 'Да':
                context.user_data['locality'] = dict(itertools.islice(context.user_data['locality'].items(),
                                                                      len(context.user_data['locality']) - 2))
                db_sess = db_session.create_session()
                reply_keyboard = [[f"{elem.title} - {get_amount(db_sess, elem, context)}"] for elem in
                                  db_sess.query(
                                      Goods
                                  ).filter(
                                      Goods.brend_id == context.user_data['add_amount']['brend_id']).all()]
                reply_keyboard.append(['Отмена'])
                update.message.reply_text(f"Выберите товар линейки",
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Нет':
                return start_menu_handler(update, context)
            else:
                update.message.reply_text(f'Ошибка. Выберите еще раз',
                                          reply_markup=ReplyKeyboardMarkup([['Да'], ['Нет']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            #
            # Наличие жидкости
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Наличие':
            db_sess = db_session.create_session()
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            elif update.message.text == 'Общее':
                text_amount = f'Общее наличие:\n'
                for elem in sorted(db_sess.query(Brends).all(), key=lambda x: -(x.price)):
                    text_amount += f'\n{elem.brend} {elem.price} рублей\n'
                    for ele in db_sess.query(Goods).filter(Goods.brend == elem).all():
                        amount_good = 0
                        for el in db_sess.query(Delivery_goods).filter(Delivery_goods.good == ele).all():
                            amount_good += el.amount
                        text_amount += f'{ele.title} {amount_good}\n'
                if len(text_amount) > 4000:
                    for i in range(len(text_amount) // 4000 + 1):
                        update.message.reply_text(text_amount[i * 4000: (i + 1) * 4000])
                else:
                    update.message.reply_text(text_amount)
                return start_menu_handler(update, context)
            elif (update.message.text,) in db_sess.query(Deliverymen.name).all():
                deliver = db_sess.query(Deliverymen).filter(Deliverymen.name == update.message.text).first()
                text_amount = f'Жидкость в наличии у {deliver.name}:\n'
                for elem in sorted(db_sess.query(Brends).all(), key=lambda x: -(x.price)):
                    text_amount += f'\n{elem.brend} {elem.price} рублей\n'
                    for el in db_sess.query(Goods).filter(Goods.brend == elem).all():
                        deliv_good = db_sess.query(Delivery_goods).filter(
                            Delivery_goods.good == el,
                            Delivery_goods.deliveryman == deliver).first()
                        text_amount += f'{el.title} {deliv_good.amount}\n'
                if len(text_amount) > 4000:
                    for i in range(len(text_amount) // 4000 + 1):
                        update.message.reply_text(text_amount[i * 4000: (i + 1) * 4000])
                else:
                    update.message.reply_text(text_amount)
                return start_menu_handler(update, context)
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Проверка 1':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif '-' in update.message.text:
                db_sess = db_session.create_session()
                sales_good = db_sess.query(Sales).filter(
                    Sales.date == datetime.date(int(update.message.text.split(' ')[0].split('-')[0]),
                                                int(update.message.text.split(' ')[0].split('-')[1]),
                                                int(update.message.text.split(' ')[0].split('-')[2])),
                    Sales.deliveryman_id == db_sess.query(Deliverymen).filter(
                        Deliverymen.name == update.message.text.split(' ')[1]).first().id).first()
                if sales_good:
                    sales_good.is_send = True
                    db_sess.add(sales_good)
                    db_sess.commit()
                    update.message.reply_text('Проверено')
                    return start_menu_handler(update, context)
                else:
                    update.message.reply_text('Ничего не найдено')
                    return start_menu_handler(update, context)
            else:
                return start_menu_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Статистика':
            db_sess = db_session.create_session()
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            elif update.message.text == 'Общее':
                context.user_data['deliveryman'] = None
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Статистика 2'
                reply_keyboard = [['Число месяца'], ['Месяц'], ['Отмена']]
                update.message.reply_text('Выберите промежуток',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif (update.message.text,) in db_sess.query(Deliverymen.name).all():
                context.user_data['deliveryman'] = db_sess.query(Deliverymen).filter(
                    Deliverymen.name == update.message.text).first().id
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Статистика 2'
                reply_keyboard = [['Число месяца'], ['Месяц'], ['Отмена']]
                update.message.reply_text('Выберите промежуток',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Статистика 2':
            db_sess = db_session.create_session()
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif update.message.text == 'Месяц':
                if context.user_data['deliveryman']:
                    salary = 0
                    total = 0
                    amount = 0
                    income = 0
                    for elem in db_sess.query(Sales).filter(
                            Sales.deliveryman_id == context.user_data['deliveryman'],
                            extract('month', Sales.date) == datetime.datetime.now().date().month,
                            extract('year', Sales.date) == datetime.datetime.now().date().year,
                            Sales.is_send == True).all():
                        salary += elem.sales_salary
                        total += elem.total
                        amount += len(elem.deliverygood_ids.split('&'))
                        if elem.income is not None:
                            income += elem.income
                    update.message.reply_text(f"Статистика "
                                              f"{db_sess.query(Deliverymen).get(context.user_data['deliveryman']).name}\n"
                                              f"Денег {calculate_money(db_sess, context.user_data['deliveryman'])}\n"
                                              f"Зарплата: {salary}\n"
                                              f"Скинуто: {total}\n"
                                              f"Продано штук: {amount}\n"
                                              f"Прибыль: {income}\n"
                                              f"Средняя цена позиции: {(total + salary) / amount if amount > 0 else 0}\n")
                    return start_menu_handler(update, context)
                else:
                    salary = 0
                    total = 0
                    amount = 0
                    income = 0
                    for elem in db_sess.query(Sales).filter(
                            extract('month', Sales.date) == datetime.datetime.now().date().month,
                            extract('year', Sales.date) == datetime.datetime.now().date().year,
                            Sales.is_send == True).all():
                        salary += elem.sales_salary
                        total += elem.total
                        amount += len(elem.deliverygood_ids.split('&'))
                        if elem.income is not None:
                            income += elem.income
                    update.message.reply_text(f"Статистика общее\n"
                                              f"Денег {calculate_money(db_sess, None)}\n"
                                              f"Зарплата: {salary}\n"
                                              f"Скинуто: {total}\n"
                                              f"Продано штук: {amount}\n"
                                              f"Прибыль: {income}\n"
                                              f"Средняя цена позиции: {(total + salary) / amount if amount > 0 else 0}\n")
                    return start_menu_handler(update, context)
            elif update.message.text == 'Число месяца':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Статистика 3'
                if context.user_data['deliveryman']:
                    reply_keyboard = [[str(elem.date)] for elem in db_sess.query(Sales).filter(
                        Sales.deliveryman_id == context.user_data['deliveryman'],
                        extract('month', Sales.date) == datetime.datetime.now().date().month,
                        extract('year', Sales.date) == datetime.datetime.now().date().year,
                        Sales.is_send == True).all()]
                    reply_keyboard.append(['Отмена'])
                    update.message.reply_text('Выберите день',
                                              reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                               resize_keyboard=True,
                                                                               one_time_keyboard=True))
                else:
                    reply_keyboard = [[str(elem.date)] for elem in db_sess.query(Sales).filter(
                        extract('month', Sales.date) == datetime.datetime.now().date().month,
                        extract('year', Sales.date) == datetime.datetime.now().date().year,
                        Sales.is_send == True).all()]
                    reply_keyboard.append(['Отмена'])
                    update.message.reply_text('Выберите день',
                                              reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                               resize_keyboard=True,
                                                                               one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Статистика 3':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif '-' in update.message.text:
                db_sess = db_session.create_session()
                if context.user_data['deliveryman']:
                    sales_good = db_sess.query(Sales).filter(
                        Sales.deliveryman_id == context.user_data['deliveryman'],
                        Sales.date == datetime.date(int(update.message.text.split('-')[0]),
                                                    int(update.message.text.split('-')[1]),
                                                    int(update.message.text.split('-')[2]))
                    ).first()
                    kolvo = len(sales_good.deliverygood_ids.split('&'))
                    sredn = (sales_good.total + sales_good.sales_salary) / kolvo if kolvo > 0 else 0
                    update.message.reply_text(f"Статитстика "
                                              f"{db_sess.query(Deliverymen).get(context.user_data['deliveryman']).name}\n"
                                              f"Зарплата: {sales_good.sales_salary}\n"
                                              f"Скинуто: {sales_good.total}\n"
                                              f"Продано штук: {kolvo}\n"
                                              f"Прибыль: {sales_good.income if sales_good.income is not None else 0}\n"
                                              f"Средняя цена позиции: {sredn}")
                    return start_menu_handler(update, context)
                else:
                    salary = 0
                    total = 0
                    amount = 0
                    income = 0
                    for elem in db_sess.query(Sales).filter(
                            Sales.date == datetime.date(int(update.message.text.split('-')[0]),
                                                        int(update.message.text.split('-')[1]),
                                                        int(update.message.text.split('-')[2]))).all():
                        salary += elem.sales_salary
                        total += elem.total
                        amount += len(elem.deliverygood_ids.split("&"))
                        if elem.income is not None:
                            income += elem.income
                    update.message.reply_text(f'Статитстика общее\n'
                                              f'Зарплата: {salary}\n'
                                              f'Скинуто: {total}\n'
                                              f'Продано штук: {amount}\n'
                                              f"Прибыль: {income}\n"
                                              f'Средняя цена позиции: {(salary + total) / amount if amount > 0 else 0}\n')
                    return start_menu_handler(update, context)
            else:
                return error_handler(update, context)
        #
        # Обработка xlsx редактирования наличия
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Выслать форму 2':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            with open("goods.xlsx", 'wb') as f:
                context.bot.get_file(update.message.document).download(out=f)
            f.close
            wb = load_workbook('goods.xlsx')
            ws = wb.active
            xlsx_deliverymen = {}
            for elem in ws[1]:
                if elem.value is not None:
                    xlsx_deliverymen[elem.column] = elem.value
            spis = []
            for elem in ws['A']:
                if elem.value is not None:
                    spis.append(elem)
            schetchik = 1
            db_sess = db_session.create_session()
            for elem in spis:
                amount_brend_zavoz = 0
                for row in ws.iter_rows(min_row=elem.row + 1,
                                        max_row=spis[schetchik].row - 1 if schetchik < len(spis) else None,
                                        min_col=elem.column + 1,
                                        max_col=elem.column + 1 + len(xlsx_deliverymen)):
                    taste = row[0]
                    for cell in row[1:]:
                        if cell.value is not None:
                            good_deliver = db_sess.query(Delivery_goods).filter(
                                Delivery_goods.good == db_sess.query(
                                    Goods
                                ).filter(
                                    Goods.title == taste.value,
                                    Goods.brend == db_sess.query(Brends).filter(
                                        Brends.brend == elem.value).first()
                                ).first(),
                                Delivery_goods.deliveryman == db_sess.query(Deliverymen).filter(
                                    Deliverymen.name == xlsx_deliverymen[cell.column]).first()).first()
                            amount_brend_zavoz += int(cell.value)
                            good_deliver.amount += int(cell.value)
                            db_sess.add(good_deliver)
                if amount_brend_zavoz > 0:
                    purch = db_sess.query(Purchase).filter(
                        Purchase.brend == db_sess.query(Brends).filter(Brends.brend == elem.value).first(),
                        Purchase.amount > 0).all()[-1]
                    purch.amount += amount_brend_zavoz
                    db_sess.add(purch)
                elif amount_brend_zavoz < 0:
                    while amount_brend_zavoz != 0:
                        purch = db_sess.query(Purchase).filter(
                            Purchase.brend == db_sess.query(Brends).filter(Brends.brend == elem.value).first(),
                            Purchase.amount > 0).first()
                        if -amount_brend_zavoz > purch.amount:
                            amount_brend_zavoz += purch.amount
                            purch.amount = 0
                        else:
                            purch.amount += amount_brend_zavoz
                            amount_brend_zavoz = 0
                        db_sess.add(purch.amount)
                        db_sess.commit()
                db_sess.commit()
                schetchik += 1
            return start_menu_handler(update, context)
        #
        # обработка xlsx нового закупа
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Новый закуп 2':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            with open("goods.xlsx", 'wb') as f:
                context.bot.get_file(update.message.document).download(out=f)
            f.close
            wb = load_workbook('goods.xlsx')
            ws = wb.active
            xlsx_deliverymen = {}
            for elem in ws[1]:
                if elem.value is not None:
                    xlsx_deliverymen[elem.column] = elem.value
            spis = []
            dict_price = {}
            for elem in ws['A']:
                if elem.value is not None and 'рублей' not in elem.value:
                    spis.append(elem)
                elif elem.value is not None and 'рублей' in elem.value:
                    dict_price[spis[-1].value] = elem.value
            schetchik = 1
            db_sess = db_session.create_session()
            for elem in spis:
                amount_brend_zavoz = 0
                for row in ws.iter_rows(min_row=elem.row + 1,
                                        max_row=spis[schetchik].row - 1 if schetchik < len(spis) else None,
                                        min_col=elem.column + 1,
                                        max_col=elem.column + 1 + len(xlsx_deliverymen)):
                    taste = row[0]
                    for cell in row[1:]:
                        if cell.value is not None:
                            good_deliver = db_sess.query(Delivery_goods).filter(
                                Delivery_goods.good == db_sess.query(
                                    Goods
                                ).filter(
                                    Goods.title == taste.value,
                                    Goods.brend == db_sess.query(Brends).filter(
                                        Brends.brend == elem.value).first()
                                ).first(),
                                Delivery_goods.deliveryman == db_sess.query(Deliverymen).filter(
                                    Deliverymen.name == xlsx_deliverymen[cell.column]).first()).first()
                            amount_brend_zavoz += int(cell.value)
                            good_deliver.amount += int(cell.value)
                            db_sess.add(good_deliver)
                if amount_brend_zavoz > 0:
                    purch = Purchase(date=datetime.datetime.now().date(),
                                     brend=db_sess.query(Brends).filter(Brends.brend == elem.value).first(),
                                     amount_purchase=amount_brend_zavoz,
                                     amount=amount_brend_zavoz,
                                     price=float(dict_price[elem.value].split(' ')[0]))
                    db_sess.add(purch)
                db_sess.commit()
                schetchik += 1
            return start_menu_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Добавить описание 2':
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            else:
                brend_id = check_brend(update.message.text)
                if brend_id:
                    context.user_data['redactor_brend'] = {0: brend_id}
                    context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Добавить описание 3'
                    update.message.reply_text('Отправьте текст с фото',
                                              reply_markup=ReplyKeyboardMarkup([['Отмена']],
                                                                               resize_keyboard=True,
                                                                               one_time_keyboard=True))
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Добавить описание 3':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            else:
                db_sess = db_session.create_session()
                brend = db_sess.query(Brends).get(context.user_data['redactor_brend'][0])
                brend.photo_link = update.message.photo[-1].get_file(
                ).download(f"photos/{brend.id}.png")
                with open(f"description/{brend.id}.txt", 'w') as f:
                    f.write(update.message.caption)
                brend.txt_file = f"description/{brend.id}.txt"
                db_sess.add(brend)
                db_sess.commit()
                return start_menu_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Описание вкусов 2':
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            else:
                brend = db_session.create_session().query(Brends
                                                          ).filter(Brends.brend == update.message.text).first()
                update.message.reply_photo(photo=open(brend.photo_link, 'rb'),
                                           caption=f"{brend.brend} {brend.price} рублей\n\n"
                                                   f"{open(brend.txt_file, 'r').read()}")
                return start_menu_handler(update, context)
    #
    # Меню доставщика
    elif update.message.chat.id in deliverymen.values():
        if context.user_data['locality'][len(context.user_data['locality'])] == 'Старт':
            if update.message.text == 'Продать':
                db_sess = db_session.create_session()
                deliver = db_sess.query(Deliverymen).get(context.user_data['user_id_db'])
                reply_keyboard = [[elem.brend] for elem in sorted(db_sess.query(Brends).all(),
                                                                  key=lambda x: -(x.price)) if
                                  get_amount_brend(elem, deliver) > 0]
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Продать 1'
                update.message.reply_text('Выберите линейку',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Наличие':
                db_sess = db_session.create_session()
                deliver = db_sess.query(Deliverymen).get(context.user_data['user_id_db'])
                text_amount = f'Жидкость в наличии у {deliver.name}:\n'
                for elem in sorted(db_sess.query(Brends).all(), key=lambda x: -(x.price)):
                    if get_amount_brend(elem, deliver) > 0:
                        text_amount += f'\n{elem.brend} {elem.price} рублей\n'
                        for el in db_sess.query(Goods).filter(Goods.brend == elem).all():
                            deliv_good = db_sess.query(Delivery_goods).filter(
                                Delivery_goods.good == el,
                                Delivery_goods.deliveryman == deliver).first()
                            if deliv_good.amount > 0:
                                text_amount += f'{el.title} {deliv_good.amount}\n'
                if len(text_amount) > 4000:
                    for i in range(len(text_amount) // 4000 + 1):
                        update.message.reply_text(text_amount[i * 4000: (i + 1) * 4000])
                else:
                    update.message.reply_text(text_amount)
                return start_menu_handler(update, context)
            elif update.message.text == 'Отправить на проверку':
                send_on_check = db_session.create_session().query(
                    Sales).filter(Sales.on_check == False,
                                  Sales.deliveryman_id == context.user_data['user_id_db']).all()
                if send_on_check:
                    context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Проверка 1'
                    reply_keyboard = [[str(elem.date) + ' ' + str(elem.total)] for elem in send_on_check]
                    reply_keyboard.append(['Отмена'])
                    update.message.reply_text('Выберите дату',
                                              reply_markup=ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True,
                                                                               one_time_keyboard=True))
                else:
                    update.message.reply_text('Ничего не найдено')
                    return start_menu_handler(update, context)
            elif update.message.text == 'Статистика за день':
                today = db_session.create_session().query(Sales).filter(
                    Sales.deliveryman_id == context.user_data['user_id_db'],
                    Sales.date == datetime.datetime.now().date()).first()
                if today:
                    sredn = (today.total + today.sales_salary) / len(today.deliverygood_ids.split("&"))
                    update.message.reply_text(f'Зарплата: {today.sales_salary}\n'
                                              f'Скинуть: {today.total}\n'
                                              f'Продано штук: {len(today.deliverygood_ids.split("&"))}\n'
                                              f'Средняя цена позиции: {sredn}\n')
                else:
                    update.message.reply_text('Пока что ничего не продано')
                return start_menu_handler(update, context)
            elif update.message.text == 'Статистика за месяц':
                month = db_session.create_session().query(Sales).filter(
                    Sales.deliveryman_id == context.user_data['user_id_db'],
                    extract('month', Sales.date) == datetime.datetime.now().date().month,
                    extract('year', Sales.date) == datetime.datetime.now().date().year,
                    Sales.is_send == True).all()
                salary = 0
                total = 0
                amount = 0
                for elem in month:
                    salary += elem.sales_salary
                    total += elem.total
                    amount += len(elem.deliverygood_ids.split("&"))
                update.message.reply_text(f'Зарплата: {salary}\n'
                                          f'Скинуто: {total}\n'
                                          f'Продано штук: {amount}\n'
                                          f'Средняя цена позиции: {(salary + total) / amount if amount > 0 else 0}\n')
                return start_menu_handler(update, context)
            elif update.message.text == 'Нужно перевести':
                need_send = db_session.create_session().query(Sales).filter(
                    Sales.deliveryman_id == context.user_data['user_id_db'],
                    Sales.is_send == False).all()
                stroka = ''
                for elem in need_send:
                    stroka += f"{elem.date}\n" \
                              f"{elem.total}\n\n"
                if stroka != '':
                    update.message.reply_text(stroka + f'Не забывай приписку долг и сегодняшнюю дату\n'
                                                       f'Сбер {sberbank}\n'
                                                       f'Альфа {alfabank}')
                else:
                    update.message.reply_text('Всё скинуто')
                return start_menu_handler(update, context)
            elif update.message.text == 'Описание вкусов':
                reply_keyboard = [[elem.brend] for elem in sorted(db_session.create_session(
                ).query(Brends).filter(Brends.photo_link != '').all(), key=lambda x: -(x.price))]
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Описание вкусов 2'
                update.message.reply_text('Выберите линейку',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Продать 1':
            brend_id = check_brend(update.message.text)
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            elif brend_id:
                db_sess = db_session.create_session()
                context.user_data['sell_good'] = {'brend_id': brend_id}
                reply_keyboard = [[f"{elem.title} - "
                                   f"{get_amount_2(db_sess, elem, context.user_data['user_id_db'])}"]
                                  for elem in db_sess.query(Goods).filter(Goods.brend_id == brend_id).all() if
                                  get_amount_2(
                                      db_sess,
                                      elem,
                                      context.user_data['user_id_db']) > 0]
                reply_keyboard.append(['Отмена'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = \
                    'Продать 2'
                update.message.reply_text('Выберите товар линейки',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                update.message.reply_text('Линейка не найдена')
                return start_menu_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Продать 2':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif (update.message.text.split(' -')[0],) in db_session.create_session().query(Goods.title).all():
                context.user_data['sell_good']['delivery_good_title'] = update.message.text.split(' -')[0]
                context.user_data['locality'][len(context.user_data['locality']) + 1] = \
                    'Продать 3'
                update.message.reply_text('Скидка',
                                          reply_markup=ReplyKeyboardMarkup([['0'], ['1'], ['2'], ['Отмена']],
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Продать 3':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif update.message.text in ['0', '1', '2']:
                db_sess = db_session.create_session()
                good_deliver = db_sess.query(Delivery_goods).filter(
                    Delivery_goods.good_id == db_sess.query(
                        Goods
                    ).filter(
                        Goods.title == context.user_data['sell_good']['delivery_good_title'],
                        Goods.brend_id == context.user_data['sell_good']['brend_id']
                    ).first().id,
                    Delivery_goods.deliveryman_id ==
                    context.user_data['user_id_db']).first()
                brend_purch = db_sess.query(Purchase
                                            ).filter(Purchase.amount > 0,
                                                     Purchase.brend_id == context.user_data['sell_good'
                                                     ]['brend_id'
                                                     ]).first()
                if sell_good(good_deliver.id, context.user_data['user_id_db'],
                             context.user_data['sell_good']['brend_id'], update.message.text, brend_purch):
                    good_deliver.amount -= 1
                    db_sess.add(good_deliver)
                    brend_purch.amount -= 1
                    db_sess.add(brend_purch)
                    db_sess.commit()
                    update.message.reply_text('Успешно!')
                    return start_menu_handler(update, context)
                else:
                    return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Проверка 1':
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif '-' in update.message.text:
                db_sess = db_session.create_session()
                sales_good = db_sess.query(Sales).filter(
                    Sales.date == datetime.date(int(update.message.text.split(' ')[0].split('-')[0]),
                                                int(update.message.text.split(' ')[0].split('-')[1]),
                                                int(update.message.text.split(' ')[0].split('-')[2])),
                    Sales.deliveryman_id == context.user_data['user_id_db']
                ).first()
                if sales_good:
                    sales_good.on_check = True
                    db_sess.add(sales_good)
                    db_sess.commit()
                    update.message.reply_text('Отправлено на проверку платежа')
                    return start_menu_handler(update, context)
                else:
                    update.message.reply_text('Ничего не найдено')
                    return start_menu_handler(update, context)
            else:
                return start_menu_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Описание вкусов 2':
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            else:
                brend = db_session.create_session().query(Brends
                                                          ).filter(Brends.brend == update.message.text).first()
                update.message.reply_photo(photo=open(brend.photo_link, 'rb'),
                                           caption=f"{brend.brend} {brend.price} рублей\n\n"
                                                   f"{open(brend.txt_file, 'r').read()}")
                return start_menu_handler(update, context)
    else:
        if context.user_data['locality'][len(context.user_data['locality'])] == 'Старт':
            if update.message.text == 'Наличие':
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Наличие 1'
                reply_keyboard = [[elem + f' {delyverymen_id[elem]}'] for elem in deliverymen.keys()]
                reply_keyboard.append(['Отмена'])
                update.message.reply_text('Выберите доставщика',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True,
                                                                           one_time_keyboard=True))
            elif update.message.text == 'Доставка':
                update.message.reply_text(text_chat)
                return start_menu_handler(update, context)
            elif update.message.text == 'Описание вкусов':
                reply_keyboard = [[elem.brend] for elem in sorted(db_session.create_session(
                ).query(Brends).filter(Brends.photo_link != '').all(), key=lambda x: -(x.price))]
                reply_keyboard.append(['Назад'])
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Описание вкусов 2'
                update.message.reply_text('Выберите линейку',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard,
                                                                           resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                return error_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Наличие 1':
            db_sess = db_session.create_session()
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif (update.message.text.split(' ')[0],) in db_sess.query(Deliverymen.name).all():
                deliver = db_sess.query(Deliverymen).filter(
                    Deliverymen.name == update.message.text.split(' ')[0]).first()
                context.user_data['user'] = deliver.id
                context.user_data['locality'][len(context.user_data['locality']) + 1] = 'Наличие 2'
                reply_keyboard = [
                    [f'{elem.brend} - {elem.price} руб ' + str(
                        get_amount_brend(elem,
                                         deliver)
                    ) + ' шт'] for elem in sorted(db_sess.query(Brends).all(),
                                                  key=lambda x: -(x.price)) if get_amount_brend(elem, deliver) > 0]
                reply_keyboard.append(['Отмена'])
                update.message.reply_text('Выберите линейку',
                                          reply_markup=ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True,
                                                                           one_time_keyboard=True))
            else:
                update.message.reply_text('Ничего не найдено')
                return start_menu_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Наличие 2':
            db_sess = db_session.create_session()
            if update.message.text == 'Отмена':
                return start_menu_handler(update, context)
            elif (update.message.text.split(' -')[0],) in db_sess.query(Brends.brend).all():
                brend_good = db_sess.query(Brends).filter(
                    Brends.brend == update.message.text.split(' -')[0]).first()
                deliver = db_sess.query(Deliverymen).get(context.user_data['user'])
                text_amount = f'Жидкость {brend_good.brend + " " + str(brend_good.price)} рублей у' \
                              f' {delyverymen_id[deliver.name]}:\n'
                for el in db_sess.query(Goods).filter(Goods.brend == brend_good).all():
                    deliv_good = db_sess.query(Delivery_goods).filter(
                        Delivery_goods.good == el,
                        Delivery_goods.deliveryman == deliver).first()
                    if deliv_good.amount > 0:
                        text_amount += f'{el.title} {deliv_good.amount}\n'
                update.message.reply_text(text_amount)
                return start_menu_handler(update, context)
            else:
                update.message.reply_text('Ничего не найдено')
                return start_menu_handler(update, context)
        elif context.user_data['locality'][len(context.user_data['locality'])] == 'Описание вкусов 2':
            if update.message.text == 'Назад':
                return start_menu_handler(update, context)
            else:
                brend = db_session.create_session().query(Brends
                                                          ).filter(Brends.brend == update.message.text).first()
                update.message.reply_photo(photo=open(brend.photo_link, 'rb'),
                                           caption=f"{brend.brend} {brend.price} рублей\n\n"
                                                   f"{open(brend.txt_file, 'r').read()}\n\n"
                                                   f"Наш канал t.me/GhostVapeKgn")
                return start_menu_handler(update, context)


def calculate_money(db_sess, id):
    money = 0
    if id is None:
        for elem in db_sess.query(Brends).all():
            amount_goods = 0
            for ele in db_sess.query(Goods).filter(Goods.brend == elem).all():
                for el in db_sess.query(Delivery_goods).filter(Delivery_goods.good == ele).all():
                    amount_goods += el.amount
            money += (elem.price - elem.salary) * amount_goods
    else:
        for elem in db_sess.query(Brends).all():
            amount_goods = 0
            for ele in db_sess.query(Goods).filter(Goods.brend == elem).all():
                for el in db_sess.query(Delivery_goods).filter(
                        Delivery_goods.good == ele,
                        Delivery_goods.deliveryman_id == id).all():
                    amount_goods += el.amount
            money += (elem.price - elem.salary) * amount_goods
    return money


def delete_brand(id):
    db_sess = db_session.create_session()
    for elem in db_sess.query(Goods).filter(Goods.brend_id == id).all():
        for el in db_sess.query(Delivery_goods).filter(Delivery_goods.good == elem).all():
            db_sess.delete(el)
        db_sess.delete(elem)
    del_brend = db_sess.query(Brends).get(id)
    db_sess.delete(del_brend)
    db_sess.commit()
    return True


def sell_good(delivery_id, deliveryman_id, brend_id, discount_number, brend_purch):
    db_sess = db_session.create_session()
    sold = db_sess.query(Sales).filter(Sales.date == datetime.datetime.now().date(),
                                       Sales.deliveryman_id == deliveryman_id).first()
    brend_delivery_good = db_sess.query(Brends).get(
        db_sess.query(Goods).get(
            db_sess.query(Delivery_goods).get(
                delivery_id).good_id).brend_id)
    if discount_number == '0':
        discount = 0
    elif discount_number == '1':
        discount = db_sess.query(Brends).get(brend_id).discount_1
    elif discount_number == '2':
        discount = db_sess.query(Brends).get(brend_id).discount_2
    if sold:
        sold.deliverygood_ids += f'&{delivery_id}.{discount}'
        sold.total += brend_delivery_good.price - brend_delivery_good.salary - discount
        sold.sales_salary += brend_delivery_good.salary
        sold.income += brend_delivery_good.price - brend_delivery_good.salary - discount - brend_purch.price
        db_sess.add(sold)
        db_sess.commit()
        return True
    else:
        db_sess.add(Sales(date=datetime.datetime.now().date(),
                          deliveryman_id=deliveryman_id,
                          deliverygood_ids=f'{delivery_id}.{discount}',
                          sales_salary=brend_delivery_good.salary,
                          total=brend_delivery_good.price - brend_delivery_good.salary - discount,
                          income=brend_delivery_good.price - brend_delivery_good.salary - discount - brend_purch.price))
        db_sess.commit()
        return True


def get_amount_brend(elem, deliver):
    db_sess = db_session.create_session()
    return sum(map(lambda kort: kort[0], db_sess.query(
        Delivery_goods.amount).filter(
        Delivery_goods.good_id.in_(
            ele.id for ele in db_sess.query(Goods).filter(Goods.brend == elem).all()),
        Delivery_goods.deliveryman == deliver
    ).all()))


def get_amount_2(db_sess, elem, id):
    return db_sess.query(Delivery_goods).filter(
        Delivery_goods.good == elem,
        Delivery_goods.deliveryman_id == id).first().amount


def get_amount(db_sess, elem, context):
    return db_sess.query(Delivery_goods).filter(
        Delivery_goods.good == elem,
        Delivery_goods.deliveryman_id == context.user_data['add_amount']['deliveryman_id']).first().amount


def redact_good_title(id, old_title, new_title):
    db_sess = db_session.create_session()
    good = db_sess.query(Goods).filter(Goods.brend_id == id, Goods.title == old_title).first()
    good.title = new_title
    db_sess.add(good)
    db_sess.commit()
    return True


def check_brend(brend):
    brend_id = db_session.create_session().query(Brends).filter(Brends.brend == brend).first()
    if brend_id is not None:
        return brend_id.id
    else:
        return False


def redact_brend_price(id, price):
    db_sess = db_session.create_session()
    brend = db_sess.query(Brends).get(id)
    brend.price = price
    db_sess.add(brend)
    db_sess.commit()
    return True


def redact_brend_title(id, title):
    db_sess = db_session.create_session()
    brend = db_sess.query(Brends).get(id)
    brend.brend = title
    db_sess.add(brend)
    db_sess.commit()
    return True


def redact_brend_salary(id, salary):
    db_sess = db_session.create_session()
    brend = db_sess.query(Brends).get(id)
    brend.salary = salary
    db_sess.add(brend)
    db_sess.commit()
    return True


def add_good(id, title):
    db_sess = db_session.create_session()
    brend = db_sess.query(Brends).filter(Brends.id == id).first()
    if (title,) not in db_sess.query(Goods.title).filter(Goods.brend_id == id).all():
        db_sess.add(Goods(title=title, brend=brend))
        db_sess.commit()
        good = db_sess.query(Goods).filter(Goods.brend == brend, Goods.title == title).first()
        for el in db_sess.query(Deliverymen).all():
            delivery_good = Delivery_goods(amount=0, good=good, deliveryman=el)
            db_sess.add(delivery_good)
            db_sess.commit()
        return True


def add_goods(title, list_goods):
    db_sess = db_session.create_session()
    brend = db_sess.query(Brends).filter(Brends.brend == title).first()
    for elem in list_goods:
        db_sess.add(Goods(title=elem, brend=brend))
        db_sess.commit()
        good = db_sess.query(Goods).filter(Goods.brend == brend, Goods.title == elem).first()
        for el in db_sess.query(Deliverymen).all():
            delivery_good = Delivery_goods(amount=0, good=good, deliveryman=el)
            db_sess.add(delivery_good)
            db_sess.commit()
    return True


def main():
    updater = Updater(TOKEN)
    dp = updater.dispatcher
    dp.add_handler(CommandHandler("start", start, run_async=True))
    dp.add_handler(MessageHandler(Filters.text & ~Filters.command | Filters.document | Filters.photo,
                                  handler, run_async=True))
    dp.add_error_handler(error_handler)
    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    db_session.global_init(f"db/goods.db")
    add_deliverymen(deliverymen)
    main()
